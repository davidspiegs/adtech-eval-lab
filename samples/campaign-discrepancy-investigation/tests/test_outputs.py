"""Verifier for campaign-discrepancy-investigation task.

Computes ground truth from raw data files and checks the agent's investigation
report against expected findings. Uses partial scoring: reward = passed / total.
"""

from __future__ import annotations

import csv
import json
import os
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from rapidfuzz import fuzz


DATA_DIR = Path("/root/data")
XLSX_OUT = Path("/root/investigation_report.xlsx")
CAMPAIGN_END = date(2026, 4, 30)
VIEWABILITY_CHANGE = date(2026, 4, 21)
TAXONOMY_DATE = date(2026, 4, 15)
EXCLUDED_DOMAINS = {"gamblingreviews.com", "pokerstrategyhub.com", "sportsodds.net"}
CPM = 12.50


# ── helpers ──────────────────────────────────────────────────────────────────

def _find_sheet(wb, target):
    if target in wb.sheetnames:
        return wb[target]
    norm = target.lower().replace(" ", "_")
    for sn in wb.sheetnames:
        if sn.lower().replace(" ", "_") == norm:
            return wb[sn]
    best_score, best = 0, None
    for sn in wb.sheetnames:
        score = fuzz.ratio(sn.lower(), target.lower())
        if score > best_score:
            best_score, best = score, sn
    return wb[best] if best and best_score >= 60 else None


def _sheet_rows(wb, target):
    ws = _find_sheet(wb, target)
    if ws is None:
        return [], []
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    while headers and headers[-1] is None:
        headers.pop()
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if any(v is not None for v in vals):
            rows.append(dict(zip(headers, vals)))
    return headers, rows


def _g(row, key, default=None):
    if key in row:
        return row[key]
    norm = key.lower().replace("_", "").replace(" ", "")
    for k, v in row.items():
        if k is not None and str(k).lower().replace("_", "").replace(" ", "") == norm:
            return v
    return default


def _ws_kv(wb, sheet_name, label):
    """Find a key-value pair in a worksheet where col A = label, col B = value."""
    ws = _find_sheet(wb, sheet_name)
    if ws is None:
        return None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a is not None and str(a).strip().lower() == label.lower():
            return ws.cell(row=r, column=2).value
    return None


# ── ground truth ─────────────────────────────────────────────────────────────

_GT_CACHE = None

def gt():
    global _GT_CACHE
    if _GT_CACHE is not None:
        return _GT_CACHE

    ad_rows = list(csv.DictReader(open(DATA_DIR / "ad_server_delivery.csv")))
    moat_rows = list(csv.DictReader(open(DATA_DIR / "moat_verification.csv")))
    qa_rows = list(csv.DictReader(open(DATA_DIR / "internal_qa_log.csv")))
    domain_rows = list(csv.DictReader(open(DATA_DIR / "domain_delivery.csv")))
    with open(DATA_DIR / "iab_taxonomy_v41.json") as f:
        tax_v41 = json.load(f)
    with open(DATA_DIR / "iab_taxonomy_v42.json") as f:
        tax_v42 = json.load(f)

    total_as = sum(int(r["impressions"]) for r in ad_rows)
    total_moat_verified = sum(int(r["verified_impressions"]) for r in moat_rows)
    total_moat_viewable = sum(int(r["viewable_impressions"]) for r in moat_rows)
    total_ivt = sum(int(r["givt_impressions"]) + int(r["sivt_impressions"]) for r in moat_rows)
    total_qa = sum(int(r["test_flag_impressions"]) for r in moat_rows)

    as_by_day = defaultdict(int)
    moat_v_by_day = defaultdict(int)
    ivt_by_day = defaultdict(int)
    qa_by_day = defaultdict(int)
    for r in ad_rows:
        as_by_day[r["date"]] += int(r["impressions"])
    for r in moat_rows:
        moat_v_by_day[r["date_utc"]] += int(r["verified_impressions"])
        ivt_by_day[r["date_utc"]] += int(r["givt_impressions"]) + int(r["sivt_impressions"])
        qa_by_day[r["date_utc"]] += int(r["test_flag_impressions"])

    tz_total = 0
    for d in as_by_day:
        expected = as_by_day[d] - ivt_by_day.get(d, 0) - qa_by_day.get(d, 0)
        actual = moat_v_by_day.get(d, 0)
        diff = expected - actual
        if diff > 0:
            tz_total += diff

    non_viewable = total_moat_verified - total_moat_viewable
    total_gap = total_as - total_moat_viewable

    # viewability
    pre = [r for r in ad_rows if date.fromisoformat(r["date"]) < VIEWABILITY_CHANGE]
    post = [r for r in ad_rows if date.fromisoformat(r["date"]) >= VIEWABILITY_CHANGE]
    pre_vr = sum(int(r["viewable_impressions"]) for r in pre) / max(1, sum(int(r["impressions"]) for r in pre))
    post_vr = sum(int(r["viewable_impressions"]) for r in post) / max(1, sum(int(r["impressions"]) for r in post))

    # domain
    excl_after_imps = sum(int(r["impressions"]) for r in domain_rows
                          if r["domain"] in EXCLUDED_DOMAINS
                          and date.fromisoformat(r["date"]) >= TAXONOMY_DATE)
    excl_after_spend = sum(float(r["spend_usd"]) for r in domain_rows
                           if r["domain"] in EXCLUDED_DOMAINS
                           and date.fromisoformat(r["date"]) >= TAXONOMY_DATE)

    _GT_CACHE = {
        "total_as": total_as,
        "total_moat_verified": total_moat_verified,
        "total_moat_viewable": total_moat_viewable,
        "total_gap": total_gap,
        "ivt": total_ivt,
        "qa": total_qa,
        "timezone": tz_total,
        "non_viewable": non_viewable,
        "pre_viewability": pre_vr,
        "post_viewability": post_vr,
        "excl_domains": EXCLUDED_DOMAINS,
        "excl_imps_after": excl_after_imps,
        "excl_spend_after": excl_after_spend,
    }
    return _GT_CACHE


# ── tests ────────────────────────────────────────────────────────────────────

class TestStructure:

    def test_workbook_sheets(self):
        """All 4 required sheets exist."""
        assert XLSX_OUT.exists(), f"Missing {XLSX_OUT}"
        wb = load_workbook(XLSX_OUT, data_only=True)
        for name in ["Impression_Reconciliation", "Viewability_Analysis",
                      "Domain_Compliance", "Investigation_Summary"]:
            assert _find_sheet(wb, name) is not None, f"Missing sheet: {name}"


class TestImpressionRecon:

    def _rows(self):
        wb = load_workbook(XLSX_OUT, data_only=True)
        _, rows = _sheet_rows(wb, "Impression_Reconciliation")
        return rows

    def test_total_gap(self):
        """Total impression gap is within 10% of ground truth."""
        rows = self._rows()
        g = gt()
        totals = [r for r in rows if str(_g(r, "category", "")).lower().strip()
                  in {"total", "grand_total", "overall", "sum", "total_gap"}]
        assert len(totals) >= 1, "No total row found in Impression_Reconciliation"
        reported_gap = float(_g(totals[0], "gap", 0) or 0)
        tol = g["total_gap"] * 0.10
        assert abs(reported_gap - g["total_gap"]) < tol, (
            f"Total gap: expected ~{g['total_gap']}, got {reported_gap}"
        )

    def test_ivt_component(self):
        """IVT component identified and within 15% of ground truth."""
        rows = self._rows()
        g = gt()
        ivt_rows = [r for r in rows
                     if any(kw in str(_g(r, "category", "")).lower()
                            for kw in ["ivt", "invalid", "fraud"])]
        assert len(ivt_rows) >= 1, "No IVT category found in reconciliation"
        ivt_gap = sum(float(_g(r, "gap", 0) or 0) for r in ivt_rows)
        tol = g["ivt"] * 0.15
        assert abs(ivt_gap - g["ivt"]) < max(tol, 5000), (
            f"IVT gap: expected ~{g['ivt']}, got {ivt_gap}"
        )

    def test_non_viewable_component(self):
        """Non-viewable component identified and within 15% of ground truth."""
        rows = self._rows()
        g = gt()
        nv_rows = [r for r in rows
                    if any(kw in str(_g(r, "category", "")).lower()
                           for kw in ["viewab", "non_view", "non-view", "not_view", "unview"])]
        assert len(nv_rows) >= 1, "No non-viewable category found"
        nv_gap = sum(float(_g(r, "gap", 0) or 0) for r in nv_rows)
        tol = g["non_viewable"] * 0.15
        assert abs(nv_gap - g["non_viewable"]) < max(tol, 10000), (
            f"Non-viewable: expected ~{g['non_viewable']}, got {nv_gap}"
        )

    def test_timezone_component(self):
        """Timezone boundary component identified and within 25% of ground truth."""
        rows = self._rows()
        g = gt()
        tz_kw = ["timezone", "time_zone", "time zone", "tz_", "utc",
                  "boundary", "eastern", "coverage", "unmeasured",
                  "month_end", "month-end", "reporting_period"]
        tz_rows = [r for r in rows
                    if any(kw in str(_g(r, "category", "")).lower() for kw in tz_kw)
                    or any(kw in str(_g(r, "explanation", "")).lower()
                           for kw in ["timezone", "time zone", "utc", "eastern",
                                      "month.end", "reporting period"])]
        assert len(tz_rows) >= 1, "No timezone/coverage category found in reconciliation"
        tz_gap = sum(float(_g(r, "gap", 0) or 0) for r in tz_rows)
        tol = g["timezone"] * 0.25
        assert abs(tz_gap - g["timezone"]) < max(tol, 5000), (
            f"Timezone gap: expected ~{g['timezone']}, got {tz_gap}"
        )

    def test_components_sum(self):
        """All gap components sum to within 5% of total gap."""
        rows = self._rows()
        g = gt()
        total_labels = {"total", "grand_total", "overall", "sum", "total_gap"}
        non_total = [r for r in rows
                     if str(_g(r, "category", "")).lower().strip() not in total_labels]
        component_sum = sum(float(_g(r, "gap", 0) or 0) for r in non_total)
        tol = g["total_gap"] * 0.05
        assert abs(component_sum - g["total_gap"]) < max(tol, 10000), (
            f"Components sum {component_sum} vs total gap {g['total_gap']}"
        )


class TestViewability:

    def test_drop_date_identified(self):
        """The viewability change date (April 21) is identified."""
        wb = load_workbook(XLSX_OUT, data_only=True)
        val = _ws_kv(wb, "Viewability_Analysis", "root_cause_date")
        assert val is not None, "root_cause_date not found"
        val_str = str(val).strip()
        assert ("2026-04-21" in val_str or "04-21" in val_str
                or "04/21" in val_str or "april 21" in val_str.lower()
                or "apr 21" in val_str.lower()), (
            f"Expected date around 2026-04-21, got '{val_str}'"
        )

    def test_placement_identified(self):
        """The below-fold placement is identified as the cause."""
        wb = load_workbook(XLSX_OUT, data_only=True)
        val = _ws_kv(wb, "Viewability_Analysis", "root_cause_placement")
        if val is None:
            val = _ws_kv(wb, "Viewability_Analysis", "root_cause_description")
        assert val is not None, "No root cause placement found"
        val_str = str(val).lower()
        assert any(kw in val_str for kw in ["pl-010", "sticky", "below_fold", "below-fold", "below fold", "320x50"]), (
            f"Expected PL-010 / below-fold / sticky, got '{val}'"
        )

    def test_daily_trend(self):
        """Daily viewability data shows a clear drop around April 21."""
        wb = load_workbook(XLSX_OUT, data_only=True)
        _, rows = _sheet_rows(wb, "Viewability_Analysis")
        day_agg = defaultdict(lambda: {"imps": 0, "viewable": 0})
        for r in rows:
            d = _g(r, "date")
            if d is None:
                continue
            d_str = str(d).strip()[:10]
            try:
                date.fromisoformat(d_str)
            except (ValueError, TypeError):
                continue
            imps = float(_g(r, "impressions", 0) or 0)
            vw = float(_g(r, "viewable_impressions", 0) or 0)
            day_agg[d_str]["imps"] += imps
            day_agg[d_str]["viewable"] += vw

        assert len(day_agg) >= 20, f"Expected >=20 daily rows, got {len(day_agg)}"
        pre_days = {d: v for d, v in day_agg.items() if d < "2026-04-21"}
        post_days = {d: v for d, v in day_agg.items() if d >= "2026-04-21"}
        if pre_days and post_days:
            pre_vr = sum(v["viewable"] for v in pre_days.values()) / max(1, sum(v["imps"] for v in pre_days.values()))
            post_vr = sum(v["viewable"] for v in post_days.values()) / max(1, sum(v["imps"] for v in post_days.values()))
            assert pre_vr > post_vr + 0.10, (
                f"Expected viewability drop: pre={pre_vr:.2f}, post={post_vr:.2f}"
            )


class TestDomainCompliance:

    def _rows(self):
        wb = load_workbook(XLSX_OUT, data_only=True)
        _, rows = _sheet_rows(wb, "Domain_Compliance")
        return rows

    def test_offending_domains(self):
        """All 3 offending domains are listed."""
        rows = self._rows()
        found = {str(_g(r, "domain", "")).lower().strip() for r in rows}
        for dom in EXCLUDED_DOMAINS:
            assert dom in found, f"Missing domain: {dom}"

    def test_wasted_impressions(self):
        """Total wasted impressions after taxonomy change within 15%."""
        g = gt()
        wb = load_workbook(XLSX_OUT, data_only=True)
        val = _ws_kv(wb, "Domain_Compliance", "total_wasted_impressions")
        if val is None:
            rows = self._rows()
            val = sum(float(_g(r, "impressions_after_issue", 0) or
                            _g(r, "impressions_after_change", 0) or 0) for r in rows)
        val = float(val)
        tol = g["excl_imps_after"] * 0.15
        assert abs(val - g["excl_imps_after"]) < max(tol, 5000), (
            f"Wasted impressions: expected ~{g['excl_imps_after']}, got {val}"
        )

    def test_taxonomy_cause(self):
        """Taxonomy re-categorization identified as root cause."""
        wb = load_workbook(XLSX_OUT, data_only=True)
        val = _ws_kv(wb, "Domain_Compliance", "root_cause")
        if val is None:
            val = _ws_kv(wb, "Domain_Compliance", "taxonomy_update_date")
        assert val is not None, "No root cause / taxonomy info found"
        val_str = str(val).lower()
        assert any(kw in val_str for kw in ["taxonomy", "recategor", "re-categor", "iab", "v4.2", "4.2", "2026-04-15", "april 15"]), (
            f"Expected taxonomy re-categorization cause, got '{val}'"
        )


class TestSummary:

    def test_summary_structure(self):
        """All 3 complaints are addressed in the summary."""
        wb = load_workbook(XLSX_OUT, data_only=True)
        ws = _find_sheet(wb, "Investigation_Summary")
        assert ws is not None, "Investigation_Summary sheet not found"
        all_text = " ".join(
            str(ws.cell(row=r, column=c).value or "").lower()
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        )
        assert any(kw in all_text for kw in ["impression", "gap", "discrep"]), "Missing impression gap finding"
        assert any(kw in all_text for kw in ["viewab", "drop", "collaps"]), "Missing viewability finding"
        assert any(kw in all_text for kw in ["domain", "brand", "gambling", "exclusion"]), "Missing domain finding"

    def test_financial_impact(self):
        """Total financial impact is present and in a reasonable range."""
        wb = load_workbook(XLSX_OUT, data_only=True)
        ws = _find_sheet(wb, "Investigation_Summary")
        assert ws is not None
        impact = None
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if a and any(kw in str(a).lower().replace(" ", "_")
                         for kw in ["total_financial", "financial_impact", "total_impact"]):
                impact = ws.cell(row=r, column=2).value
                break
        assert impact is not None, "No total financial impact label found in Investigation_Summary"
        val = float(impact)
        assert val > 1000, f"Financial impact {val} seems too low (expected >$1000)"
        assert val < 500000, f"Financial impact {val} seems too high (expected <$500K)"
