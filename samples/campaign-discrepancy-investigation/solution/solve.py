#!/usr/bin/env python3
"""Oracle solver: investigates 3 campaign complaints, produces investigation_report.xlsx."""

import csv
import json
from collections import defaultdict
from datetime import date

from openpyxl import Workbook

DATA = "/root/data"
CPM = 12.50
CAMPAIGN_END = date(2026, 4, 30)
VIEWABILITY_CHANGE = date(2026, 4, 21)
TAXONOMY_DATE = date(2026, 4, 15)
EXCLUDED_DOMAINS = ["gamblingreviews.com", "pokerstrategyhub.com", "sportsodds.net"]


def read_csv(path):
    with open(path) as f:
        return list(csv.DictReader(f))


def main():
    ad_rows = read_csv(f"{DATA}/ad_server_delivery.csv")
    moat_rows = read_csv(f"{DATA}/moat_verification.csv")
    qa_rows = read_csv(f"{DATA}/internal_qa_log.csv")
    domain_rows = read_csv(f"{DATA}/domain_delivery.csv")
    placements = {r["placement_id"]: r for r in read_csv(f"{DATA}/placement_specs.csv")}
    with open(f"{DATA}/iab_taxonomy_v41.json") as f:
        tax_v41 = json.load(f)
    with open(f"{DATA}/iab_taxonomy_v42.json") as f:
        tax_v42 = json.load(f)
    with open(f"{DATA}/campaign_config.json") as f:
        config = json.load(f)

    # ── Complaint 1: Impression gap breakdown ────────────────────────────────
    total_as = sum(int(r["impressions"]) for r in ad_rows)
    total_moat_verified = sum(int(r["verified_impressions"]) for r in moat_rows)
    total_moat_viewable = sum(int(r["viewable_impressions"]) for r in moat_rows)

    total_ivt = sum(int(r["givt_impressions"]) + int(r["sivt_impressions"]) for r in moat_rows)
    total_qa = sum(int(r["test_flag_impressions"]) for r in moat_rows)

    as_by_day = defaultdict(int)
    moat_verified_by_day = defaultdict(int)
    for r in ad_rows:
        as_by_day[r["date"]] += int(r["impressions"])
    for r in moat_rows:
        moat_verified_by_day[r["date_utc"]] += int(r["verified_impressions"])

    tz_gap_by_day = {}
    for d in sorted(as_by_day.keys()):
        expected_verified = as_by_day[d] - sum(
            int(r["givt_impressions"]) + int(r["sivt_impressions"]) + int(r["test_flag_impressions"])
            for r in moat_rows if r["date_utc"] == d
        )
        actual_verified = moat_verified_by_day.get(d, 0)
        tz_gap_by_day[d] = expected_verified - actual_verified

    timezone_total = sum(max(0, v) for v in tz_gap_by_day.values())
    non_viewable = total_moat_verified - total_moat_viewable
    total_gap = total_as - total_moat_viewable

    categories = [
        ("ivt_filtered", total_as, total_as - total_ivt, total_ivt,
         "IVT (GIVT + SIVT) filtered by Moat verification but not deducted by ad server"),
        ("non_viewable", total_moat_verified, total_moat_viewable, non_viewable,
         "Non-viewable impressions — Moat counts viewable completions only"),
        ("timezone_boundary", 0, 0, timezone_total,
         "Timezone offset: ad server uses US/Eastern, Moat uses UTC — April 30 evening impressions attributed to May 1 by Moat"),
        ("test_traffic", total_as, total_as - total_qa, total_qa,
         "Internal QA/test traffic filtered by Moat test detection"),
    ]
    residual = total_gap - sum(c[3] for c in categories)

    # ── Complaint 2: Viewability analysis ────────────────────────────────────
    view_daily = defaultdict(lambda: defaultdict(lambda: {"imps": 0, "viewable": 0}))
    for r in ad_rows:
        view_daily[r["date"]][r["placement_id"]]["imps"] += int(r["impressions"])
        view_daily[r["date"]][r["placement_id"]]["viewable"] += int(r["viewable_impressions"])

    # ── Complaint 3: Domain compliance ───────────────────────────────────────
    domain_data = defaultdict(lambda: {"before": 0, "after": 0, "spend_before": 0.0, "spend_after": 0.0})
    for r in domain_rows:
        if r["domain"] in EXCLUDED_DOMAINS:
            d = date.fromisoformat(r["date"])
            if d >= TAXONOMY_DATE:
                domain_data[r["domain"]]["after"] += int(r["impressions"])
                domain_data[r["domain"]]["spend_after"] += float(r["spend_usd"])
            else:
                domain_data[r["domain"]]["before"] += int(r["impressions"])
                domain_data[r["domain"]]["spend_before"] += float(r["spend_usd"])

    total_wasted_imps = sum(v["after"] for v in domain_data.values())
    total_wasted_spend = sum(v["spend_after"] for v in domain_data.values())

    # ── Write workbook ───────────────────────────────────────────────────────
    wb = Workbook()

    # Sheet 1: Impression Reconciliation
    ws1 = wb.active
    ws1.title = "Impression_Reconciliation"
    ws1.append(["category", "ad_server_count", "moat_count", "gap", "pct_of_total_gap", "explanation"])
    for name, as_ct, moat_ct, gap_val, explanation in categories:
        ws1.append([name, as_ct if as_ct else "", moat_ct if moat_ct else "",
                     gap_val, round(gap_val / total_gap * 100, 1) if total_gap else 0, explanation])
    if abs(residual) > 100:
        ws1.append(["residual_rounding", "", "", residual,
                     round(residual / total_gap * 100, 1), "Rounding and measurement timing differences"])
    ws1.append(["total", total_as, total_moat_viewable, total_gap, 100.0, "Total impression gap"])

    # Sheet 2: Viewability Analysis
    ws2 = wb.create_sheet("Viewability_Analysis")
    ws2.append(["date", "placement_id", "impressions", "viewable_impressions", "viewability_pct"])
    for d in sorted(view_daily.keys()):
        for pl_id in sorted(view_daily[d].keys()):
            v = view_daily[d][pl_id]
            vr = round(v["viewable"] / v["imps"] * 100, 1) if v["imps"] > 0 else 0
            ws2.append([d, pl_id, v["imps"], v["viewable"], vr])

    next_row = ws2.max_row + 2
    ws2.cell(row=next_row, column=1, value="root_cause_date")
    ws2.cell(row=next_row, column=2, value=VIEWABILITY_CHANGE.isoformat())
    next_row += 1
    ws2.cell(row=next_row, column=1, value="root_cause_placement")
    ws2.cell(row=next_row, column=2, value="PL-010")
    next_row += 1
    ws2.cell(row=next_row, column=1, value="root_cause_description")
    ws2.cell(row=next_row, column=2, value="Below-fold sticky placement (Homepage_Sticky_320x50) launched on 2026-04-21 with 60% traffic allocation, viewability benchmark 18% vs ~72% above-fold average")

    # Sheet 3: Domain Compliance
    ws3 = wb.create_sheet("Domain_Compliance")
    ws3.append(["domain", "category_at_serving", "category_current",
                "on_exclusion_list", "impressions_before_issue", "impressions_after_issue",
                "wasted_spend_usd"])
    for dom in sorted(domain_data.keys()):
        v41_cat = tax_v41["mappings"].get(dom, {}).get("primary_category", "")
        v42_cat = tax_v42["mappings"].get(dom, {}).get("primary_category", "")
        ws3.append([dom, v41_cat, v42_cat, "yes",
                     domain_data[dom]["before"], domain_data[dom]["after"],
                     round(domain_data[dom]["spend_after"], 2)])

    next_row = ws3.max_row + 2
    for label, val in [
        ("root_cause", "IAB taxonomy re-categorization: domains moved from IAB9 (Hobbies) to IAB9-27 (Gambling) in v4.2 on 2026-04-15, but ad server used cached v4.1"),
        ("issue_start_date", TAXONOMY_DATE.isoformat()),
        ("total_wasted_impressions", total_wasted_imps),
        ("total_wasted_spend_usd", round(total_wasted_spend, 2)),
    ]:
        ws3.cell(row=next_row, column=1, value=label)
        ws3.cell(row=next_row, column=2, value=val)
        next_row += 1

    # Sheet 4: Investigation Summary
    ws4 = wb.create_sheet("Investigation_Summary")
    imp_financial = round(total_gap * CPM / 1000, 2)
    domain_financial = round(total_wasted_spend, 2)

    ws4.append(["finding", "impression_gap", total_gap, round(imp_financial, 2)])
    ws4.append(["finding", "viewability_drop", "below_fold_placement_PL-010", VIEWABILITY_CHANGE.isoformat()])
    ws4.append(["finding", "domain_violation", len(domain_data), total_wasted_imps])
    ws4.append([None, None, None, None])
    ws4.append(["total_financial_impact_usd", round(imp_financial + domain_financial, 2), None, None])
    ws4.append([None, None, None, None])
    ws4.append(["rank", "concern", "financial_impact_usd", "evidence_source"])
    ws4.append([1, "Impression gap: non-viewable + IVT + test traffic + timezone", round(imp_financial, 2), "Impression_Reconciliation"])
    ws4.append([2, "Domain brand safety violation on gambling sites", round(domain_financial, 2), "Domain_Compliance"])
    ws4.append([3, "Viewability collapse from below-fold placement launch", 0, "Viewability_Analysis"])

    wb.save("/root/investigation_report.xlsx")
    print(f"saved=/root/investigation_report.xlsx")


if __name__ == "__main__":
    main()
