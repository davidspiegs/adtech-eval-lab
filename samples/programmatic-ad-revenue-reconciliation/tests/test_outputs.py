"""Deterministic verifier for programmatic-ad-revenue-reconciliation 5-sheet Excel output.

Computes ground truth from input data using the same reconciliation logic the agent
must implement: weighted CPM, IVT deductions, fuzzy name matching, dual-threshold
flagging, PMP floor exception, pacing / make-good analysis, SSP yield metrics,
month-over-month variance, and executive summary roll-up.

Reads SSP billing files directly from /root/data/ssp_billing/, including the
static Index Exchange PDF that is visible to the agent.
"""

from __future__ import annotations

import csv
import json
import re

import pdfplumber
from collections import defaultdict
from datetime import date
from pathlib import Path

from rapidfuzz import fuzz
from openpyxl import load_workbook


# ── constants ────────────────────────────────────────────────────────────────

DATA_DIR = Path("/root/data")
XLSX_OUT = Path("/root/audit_report.xlsx")

FUZZY_THRESHOLD = 80
MONTH_START = date(2026, 4, 1)
MONTH_END = date(2026, 4, 30)

SSP_CANONICAL = {
    "adflow": "AdFlow Exchange",
    "magnite": "Magnite",
    "indexexchange": "Index Exchange",
    "pubmatic": "Pubmatic",
}

SSP_FILE_MAP = {
    "adflow_exchange": "adflow",
    "magnite": "magnite",
    "indexexchange": "indexexchange",
    "pubmatic": "pubmatic",
}


# ── helpers ──────────────────────────────────────────────────────────────────

def _resolve_name(ssp_id, ssp_au_name, exact_map, all_names):
    """Resolve SSP ad-unit name -> (ad_unit_id, active) via exact then fuzzy."""
    result = exact_map.get((ssp_id, ssp_au_name))
    if result:
        return result
    result = exact_map.get(("_canonical", ssp_au_name))
    if result:
        return result
    best_score, best_match = 0, None
    for sid, name, au_id, active in all_names:
        if sid != ssp_id and sid != "_canonical":
            continue
        score = fuzz.ratio(ssp_au_name.lower(), name.lower())
        if score > best_score:
            best_score = score
            best_match = (au_id, active)
    return best_match if best_score >= FUZZY_THRESHOLD else None


def _find_sheet(wb, target):
    """Find worksheet by name: exact -> normalised -> fuzzy fallback."""
    if target in wb.sheetnames:
        return wb[target]
    norm = target.lower().replace(" ", "_")
    for sn in wb.sheetnames:
        if sn.lower().replace(" ", "_") == norm:
            return wb[sn]
    best_score, best = 0, None
    for sn in wb.sheetnames:
        score = fuzz.ratio(sn.lower(), target.lower())
        if score > best_score:
            best_score, best = score, sn
    return wb[best] if best and best_score >= 60 else None


def _sheet_rows(wb, target):
    """Return (headers, list-of-dicts) for a data sheet."""
    ws = _find_sheet(wb, target)
    if ws is None:
        return [], []
    max_col = ws.max_column
    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    while headers and headers[-1] is None:
        headers.pop()
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if any(v is not None for v in vals):
            rows.append(dict(zip(headers, vals)))
    return headers, rows


def _g(row, key, default=None):
    """Flexible dict getter -- normalises key before lookup."""
    if key in row:
        return row[key]
    norm = key.lower().replace("_", "").replace(" ", "")
    for k, v in row.items():
        if k is not None and str(k).lower().replace("_", "").replace(" ", "") == norm:
            return v
    return default


def _status(row):
    """Extract and normalise a status value from a row."""
    return str(_g(row, "status", "")).strip().lower()


# ── ground-truth computation ─────────────────────────────────────────────────

def _read_pdf_billing(path):
    """Extract pipe-delimited billing rows from the static Index Exchange PDF."""
    rows = []
    cols = [
        "ssp_line_id",
        "month",
        "deal_id",
        "ssp_ad_unit_name",
        "billed_impressions",
        "gross_revenue_usd",
        "ssp_fee_pct",
        "ssp_fee_usd",
        "net_revenue_usd",
    ]
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.split("\n"):
                if "|" not in line:
                    continue
                parts = [p.strip() for p in line.split("|")]
                if len(parts) < 5 or parts[0].startswith("-"):
                    continue
                if re.match(r"[A-Z]+-LINE-\d+", parts[0]):
                    rows.append({c: parts[i] if i < len(parts) else "" for i, c in enumerate(cols)})
    return rows


def _load_ground_truth():                                      # noqa: C901
    """Recompute all ground-truth values from input data files."""

    ad_server = list(csv.DictReader(open(DATA_DIR / "ad_server_delivery.csv")))
    ivt_data = list(csv.DictReader(open(DATA_DIR / "ivt_report.csv")))
    rate_card = list(csv.DictReader(open(DATA_DIR / "deal_rate_card.csv")))
    metadata = list(csv.DictReader(open(DATA_DIR / "ad_unit_metadata.csv")))
    contract_goals = list(csv.DictReader(open(DATA_DIR / "deal_contract_goals.csv")))
    request_log = list(csv.DictReader(open(DATA_DIR / "ssp_request_log.csv")))
    historical = json.loads((DATA_DIR / "historical_metrics.json").read_text())

    ssp_billing_dir = DATA_DIR / "ssp_billing"
    ssp_rows = []
    for path in sorted(ssp_billing_dir.iterdir()):
        if path.suffix not in {".csv", ".pdf"}:
            continue
        ssp_id = None
        for prefix, sid in SSP_FILE_MAP.items():
            if path.name.startswith(prefix):
                ssp_id = sid
                break
        if ssp_id is None:
            ssp_id = path.stem.split("_")[0].lower()
        rows = (
            list(csv.DictReader(open(path)))
            if path.suffix == ".csv"
            else _read_pdf_billing(path)
        )
        for r in rows:
            r["_ssp_id"] = ssp_id
        ssp_rows.extend(rows)

    known_deals = {rc["deal_id"] for rc in rate_card}

    rc_lookup, rc_type_lookup = {}, {}
    for rc in rate_card:
        rc_lookup[(rc["deal_id"], rc["ad_unit_id"], rc["device"], rc["geo"])] = float(rc["cpm_usd"])
        rc_type_lookup[(rc["deal_id"], rc["ad_unit_id"])] = rc["rate_type"]

    del_dpdg = defaultdict(int)
    for row in ad_server:
        del_dpdg[(row["deal_id"], row["ad_unit_id"], row["device"], row["geo"])] += int(
            row["delivered_impressions"]
        )
    del_dp = defaultdict(int)
    for (d, au, dev, geo), v in del_dpdg.items():
        del_dp[(d, au)] += v

    ivt_dp = defaultdict(int)
    for row in ivt_data:
        ivt_dp[(row["deal_id"], row["ad_unit_id"])] += int(row["ivt_impressions"])

    ground_truth = {}
    for dp_key in sorted(del_dp.keys()):
        if dp_key not in rc_type_lookup:
            continue
        deal_id, ad_unit_id = dp_key
        total_del = del_dp[dp_key]
        total_ivt = ivt_dp.get(dp_key, 0)
        billable = total_del - total_ivt
        weighted_sum = sum(
            del_dpdg[(d, au, dev, geo)] * rc_lookup.get((d, au, dev, geo), 0)
            for (d, au, dev, geo) in del_dpdg
            if d == deal_id and au == ad_unit_id
        )
        w_avg_cpm = weighted_sum / total_del if total_del > 0 else 0.0
        expected_rev = round(billable / 1000.0 * w_avg_cpm, 2)
        ground_truth[dp_key] = {
            "delivered": total_del,
            "ivt": total_ivt,
            "billable": billable,
            "w_avg_cpm": w_avg_cpm,
            "expected_revenue": expected_rev,
            "rate_type": rc_type_lookup[dp_key],
        }

    exact_map, all_names = {}, []
    for m in metadata:
        for sid in SSP_CANONICAL:
            col = f"{sid}_name"
            if col in m and m[col]:
                exact_map[(sid, m[col])] = (m["ad_unit_id"], m["active"] == "true")
                all_names.append((sid, m[col], m["ad_unit_id"], m["active"] == "true"))
        exact_map[("_canonical", m["canonical_name"])] = (m["ad_unit_id"], m["active"] == "true")
        all_names.append(("_canonical", m["canonical_name"], m["ad_unit_id"], m["active"] == "true"))

    # ── reconciliation ───────────────────────────────────────────────────────
    total_expected = round(sum(gt["expected_revenue"] for gt in ground_truth.values()), 2)
    total_ssp = round(sum(float(r["net_revenue_usd"]) for r in ssp_rows), 2)

    matched_dp, flagged_ids, held_line_ids = set(), set(), set()
    ssp_for_dp, ssp_billed_for_dp = {}, {}

    for sr in ssp_rows:
        ssp_id = sr["_ssp_id"]
        deal_id = sr["deal_id"]
        ssp_au_name = sr["ssp_ad_unit_name"]

        if deal_id not in known_deals:
            held_line_ids.add(sr["ssp_line_id"])
            continue
        au_info = _resolve_name(ssp_id, ssp_au_name, exact_map, all_names)
        if au_info is None:
            held_line_ids.add(sr["ssp_line_id"])
            continue
        au_id, is_active = au_info
        if not is_active:
            held_line_ids.add(sr["ssp_line_id"])
            continue
        dp_key = (deal_id, au_id)
        if dp_key not in ground_truth:
            held_line_ids.add(sr["ssp_line_id"])
            continue

        matched_dp.add(dp_key)
        ssp_for_dp[dp_key] = ssp_id
        ssp_billed_for_dp[dp_key] = (int(sr["billed_impressions"]), float(sr["net_revenue_usd"]))

        gt_entry = ground_truth[dp_key]
        disc = round(gt_entry["expected_revenue"] - float(sr["net_revenue_usd"]), 2)
        pct = (
            round(abs(disc) / gt_entry["expected_revenue"] * 100, 2)
            if gt_entry["expected_revenue"] > 0
            else 0.0
        )
        is_pmp_above_floor = gt_entry["rate_type"] == "floor" and disc < 0
        if not is_pmp_above_floor and abs(disc) > 5.0 and pct > 3.0:
            flagged_ids.add(dp_key)

    total_adjusted = round(sum(ground_truth[dp]["expected_revenue"] for dp in matched_dp), 2)

    # ── pacing & make-goods ──────────────────────────────────────────────────
    pacing_rows = []
    total_makegood_dollars = 0.0

    for cg in contract_goals:
        dp_key = (cg["deal_id"], cg["ad_unit_id"])
        contracted = int(cg["contracted_impressions"])
        start_d = date.fromisoformat(cg["contract_start"])
        end_d = date.fromisoformat(cg["contract_end"])
        elapsed_start = max(start_d, MONTH_START)
        elapsed_end = min(end_d, MONTH_END)
        days_elapsed = (elapsed_end - elapsed_start).days + 1
        days_total = (end_d - start_d).days + 1

        delivered = ground_truth.get(dp_key, {}).get("delivered", 0)
        projected_eom = delivered * days_total / days_elapsed if days_elapsed > 0 else 0
        delivery_pct = delivered / contracted * 100 if contracted > 0 else 0
        makegood_imps = max(0, contracted - projected_eom)
        w_cpm = ground_truth.get(dp_key, {}).get("w_avg_cpm", 0)
        mg_dollars = makegood_imps / 1000.0 * w_cpm
        total_makegood_dollars += mg_dollars

        expected_pct = days_elapsed / days_total * 100
        if delivery_pct >= 100:
            status = "complete"
        elif delivery_pct > expected_pct * 1.05:
            status = "ahead"
        elif delivery_pct >= expected_pct * 0.95:
            status = "on_track"
        elif delivery_pct >= expected_pct * 0.85:
            status = "at_risk"
        else:
            status = "behind"

        pacing_rows.append(
            {
                "deal_id": cg["deal_id"],
                "ad_unit_id": cg["ad_unit_id"],
                "contracted": contracted,
                "days_elapsed": days_elapsed,
                "days_total": days_total,
                "delivered": delivered,
                "projected_eom": projected_eom,
                "delivery_pct": delivery_pct,
                "makegood_imps": makegood_imps,
                "makegood_dollars": mg_dollars,
                "status": status,
            }
        )

    # ── SSP yield ────────────────────────────────────────────────────────────
    ssp_requests = defaultdict(int)
    for row in request_log:
        ssp_requests[row["ssp_id"]] += int(row["ad_requests"])

    ssp_imps = defaultdict(int)
    ssp_rev = defaultdict(float)
    for sr in ssp_rows:
        ssp_imps[sr["_ssp_id"]] += int(sr["billed_impressions"])
        ssp_rev[sr["_ssp_id"]] += float(sr["net_revenue_usd"])

    ssp_flagged_ct = defaultdict(int)
    ssp_matched_ct = defaultdict(int)
    for dp_key in matched_dp:
        sid = ssp_for_dp[dp_key]
        ssp_matched_ct[sid] += 1
        if dp_key in flagged_ids:
            ssp_flagged_ct[sid] += 1

    yield_rows = []
    for ssp_id in SSP_CANONICAL:
        ad_req = ssp_requests.get(ssp_id, 0)
        imps = ssp_imps.get(ssp_id, 0)
        rev = ssp_rev.get(ssp_id, 0)
        ecpm = rev / imps * 1000 if imps > 0 else 0
        matched = ssp_matched_ct.get(ssp_id, 0)
        flagged = ssp_flagged_ct.get(ssp_id, 0)
        disc_rate = flagged / matched if matched > 0 else 0
        yield_rows.append(
            {
                "ssp_id": ssp_id,
                "ad_requests": ad_req,
                "impressions_filled": imps,
                "fill_rate": imps / ad_req if ad_req > 0 else 0,
                "total_net_revenue": rev,
                "ecpm": ecpm,
                "discrepancy_rate": disc_rate,
            }
        )
    yield_rows.sort(key=lambda r: r["ecpm"], reverse=True)
    for i, yr in enumerate(yield_rows):
        yr["rank"] = i + 1

    # ── variance vs prior month ──────────────────────────────────────────────
    total_imps_all = sum(ssp_imps.values())
    total_delivered = sum(g["delivered"] for g in ground_truth.values())
    total_ivt = sum(g["ivt"] for g in ground_truth.values())
    current_ivt_rate = total_ivt / total_delivered * 100 if total_delivered > 0 else 0
    current_disc_rate = len(flagged_ids) / len(matched_dp) * 100 if matched_dp else 0
    current_ecpm = sum(ssp_rev.values()) / total_imps_all * 1000 if total_imps_all > 0 else 0

    variance_metrics = {
        "total_net_revenue_usd": {"current": total_ssp, "prior": historical["total_net_revenue_usd"]},
        "ivt_rate_pct": {"current": current_ivt_rate, "prior": historical["ivt_rate_pct"]},
        "discrepancy_rate_pct": {"current": current_disc_rate, "prior": historical["discrepancy_rate_pct"]},
        "held_row_count": {"current": float(len(held_line_ids)), "prior": float(historical["held_row_count"])},
        "avg_ecpm_usd": {"current": current_ecpm, "prior": historical["avg_ecpm_usd"]},
    }
    for vals in variance_metrics.values():
        p, c = vals["prior"], vals["current"]
        vals["change_pct"] = (c - p) / p * 100 if p != 0 else 0
        acp = abs(vals["change_pct"])
        vals["status"] = "normal" if acp <= 10 else ("watch" if acp <= 25 else "alert")

    # ── executive summary ────────────────────────────────────────────────────
    flagged_rev_at_risk = sum(
        abs(ground_truth[dp]["expected_revenue"] - ssp_billed_for_dp[dp][1])
        for dp in flagged_ids
        if dp in ssp_billed_for_dp
    )
    num_flagged = len(flagged_ids)
    if num_flagged >= 3 or total_makegood_dollars > 2000:
        overall_health = "Red"
    elif num_flagged >= 1 or total_makegood_dollars >= 500:
        overall_health = "Yellow"
    else:
        overall_health = "Green"

    return {
        "ground_truth": ground_truth,
        "total_expected": total_expected,
        "total_ssp": total_ssp,
        "total_adjusted": total_adjusted,
        "flagged_ids": flagged_ids,
        "held_line_ids": held_line_ids,
        "matched_dp": matched_dp,
        "ssp_for_dp": ssp_for_dp,
        "ssp_billed_for_dp": ssp_billed_for_dp,
        "pacing_rows": pacing_rows,
        "total_makegood_dollars": total_makegood_dollars,
        "yield_rows": yield_rows,
        "ssp_requests": ssp_requests,
        "variance_metrics": variance_metrics,
        "historical": historical,
        "overall_health": overall_health,
        "flagged_rev_at_risk": flagged_rev_at_risk,
    }


_GT_CACHE = None


def gt():
    global _GT_CACHE
    if _GT_CACHE is None:
        _GT_CACHE = _load_ground_truth()
    return _GT_CACHE


# ── test classes (15 tests) ──────────────────────────────────────────────────


class TestWorkbookStructure:

    def test_workbook_structure(self):
        """File exists with all five required sheets."""
        assert XLSX_OUT.exists(), f"Missing {XLSX_OUT}"
        wb = load_workbook(XLSX_OUT, data_only=True)
        assert len(wb.sheetnames) >= 5, f"Expected >=5 sheets, got {wb.sheetnames}"
        for name in [
            "Revenue_Reconciliation",
            "Pacing_And_MakeGoods",
            "SSP_Yield_Analysis",
            "Variance_vs_Prior_Month",
            "Executive_Summary",
        ]:
            assert _find_sheet(wb, name) is not None, f"Missing sheet: {name}"


class TestReconciliation:

    def _rows(self):
        wb = load_workbook(XLSX_OUT, data_only=True)
        _, rows = _sheet_rows(wb, "Revenue_Reconciliation")
        return rows

    def test_recon_row_count(self):
        """Row count matches matched + held, with tolerance for optional unbilled rows."""
        rows = self._rows()
        g = gt()
        expected = len(g["matched_dp"]) + len(g["held_line_ids"])
        assert expected <= len(rows) <= expected + 5, (
            f"Expected {expected}-{expected+5} rows, got {len(rows)}"
        )

    def test_recon_flagged_count(self):
        """Flagged row count is at least the expected count (extras tolerated for unbilled)."""
        rows = self._rows()
        g = gt()
        flagged = [r for r in rows if _status(r).startswith("flagged")]
        min_expected = len(g["flagged_ids"])
        assert min_expected <= len(flagged) <= min_expected + 4, (
            f"Expected {min_expected}-{min_expected+4} flagged, got {len(flagged)}"
        )

    def test_recon_held_and_zeroed(self):
        """Held row count is exact and held rows have zeroed publisher-side fields."""
        rows = self._rows()
        g = gt()
        held = [r for r in rows if _status(r) == "held_unmatched"]
        assert len(held) == len(g["held_line_ids"]), (
            f"Expected {len(g['held_line_ids'])} held, got {len(held)}"
        )
        for r in held:
            delivered = float(_g(r, "delivered_impressions", 0) or 0)
            assert delivered == 0, (
                f"Held row should have delivered_impressions=0, got {delivered}"
            )

    def test_recon_total_expected_revenue(self):
        """Sum of expected_revenue_usd matches total_adjusted or total_expected."""
        rows = self._rows()
        g = gt()
        total = sum(float(_g(r, "expected_revenue_usd", 0) or 0) for r in rows)
        within_adjusted = abs(total - g["total_adjusted"]) < 1.0
        within_expected = abs(total - g["total_expected"]) < 1.0
        assert within_adjusted or within_expected, (
            f"Sheet total {total:.2f} not close to adjusted {g['total_adjusted']:.2f} "
            f"or expected {g['total_expected']:.2f}"
        )

    def test_recon_billable_spot_check(self):
        """Spot-check billable_impressions for up to 5 matched deal/ad_unit pairs."""
        rows = self._rows()
        g = gt()
        row_map = {}
        for r in rows:
            dp = (str(_g(r, "deal_id", "")), str(_g(r, "ad_unit_id", "")))
            row_map[dp] = r
        checked = 0
        for dp_key in sorted(g["matched_dp"]):
            if dp_key in row_map:
                expected = g["ground_truth"][dp_key]["billable"]
                actual = int(float(_g(row_map[dp_key], "billable_impressions", 0) or 0))
                assert actual == expected, f"{dp_key}: billable {expected} vs {actual}"
                checked += 1
                if checked >= 5:
                    break
        assert checked > 0, "No matched rows found for spot check"

    def test_recon_flagged_deal_ids(self):
        """All expected flagged deal/ad_unit pairs are present (extras tolerated)."""
        rows = self._rows()
        g = gt()
        submitted = {
            (str(_g(r, "deal_id")), str(_g(r, "ad_unit_id")))
            for r in rows
            if _status(r).startswith("flagged")
        }
        assert g["flagged_ids"].issubset(submitted), (
            f"Missing flagged IDs: {g['flagged_ids'] - submitted}"
        )

    def test_recon_valid_statuses(self):
        """Every row has a recognised status value."""
        rows = self._rows()
        valid = {
            "reconciled", "reconciled_above_floor",
            "flagged_underbilled", "flagged_overbilled", "held_unmatched",
            "flagged_unbilled", "unbilled",
        }
        for r in rows:
            s = _status(r)
            assert s in valid, f"Invalid status '{s}'"


class TestPacing:

    def _rows(self):
        wb = load_workbook(XLSX_OUT, data_only=True)
        _, rows = _sheet_rows(wb, "Pacing_And_MakeGoods")
        return rows

    def test_pacing_structure(self):
        """Correct row count, valid statuses, and at least one behind/at_risk deal."""
        rows = self._rows()
        g = gt()
        assert len(rows) == len(g["pacing_rows"]), (
            f"Expected {len(g['pacing_rows'])} pacing rows, got {len(rows)}"
        )
        valid = {"complete", "ahead", "on_track", "at_risk", "behind"}
        statuses = set()
        for r in rows:
            s = str(_g(r, "status", "")).lower()
            assert s in valid, f"Invalid pacing status '{s}'"
            statuses.add(s)
        assert statuses & {"behind", "at_risk"}, (
            f"Expected at least one behind/at_risk, got {statuses}"
        )

    def test_pacing_values(self):
        """Make-good logic, delivery percentage range, and contracted impressions match."""
        rows = self._rows()
        g = gt()
        gt_map = {(p["deal_id"], p["ad_unit_id"]): p["contracted"] for p in g["pacing_rows"]}
        for r in rows:
            pct = float(_g(r, "delivery_pct", 0) or 0)
            assert 0 <= pct <= 200, f"delivery_pct {pct} out of range"
            if str(_g(r, "status", "")).lower() == "behind":
                mg = float(_g(r, "makegood_dollars", 0) or 0)
                assert mg > 0, (
                    f"Behind deal {_g(r, 'deal_id')} should have makegood_dollars > 0"
                )
            dp = (str(_g(r, "deal_id", "")), str(_g(r, "ad_unit_id", "")))
            if dp in gt_map:
                actual = int(float(_g(r, "contracted_impressions", 0) or 0))
                assert actual == gt_map[dp], f"{dp}: contracted {gt_map[dp]} vs {actual}"


class TestYield:

    def _rows(self):
        wb = load_workbook(XLSX_OUT, data_only=True)
        _, rows = _sheet_rows(wb, "SSP_Yield_Analysis")
        return rows

    def test_yield_structure_and_ranking(self):
        """4 SSP rows, ranked correctly by eCPM descending, all with positive ad requests."""
        rows = self._rows()
        assert len(rows) == 4, f"Expected 4 SSP rows, got {len(rows)}"
        ecpms, ranks = [], []
        for r in rows:
            ecpms.append(float(_g(r, "ecpm", 0) or 0))
            ranks.append(int(float(_g(r, "ranked_position", 0) or 0)))
            ar = int(float(_g(r, "ad_requests", 0) or 0))
            assert ar > 0, f"ad_requests should be > 0 for {_g(r, 'ssp')}"
        assert sorted(ranks) == [1, 2, 3, 4], f"Ranks should be 1-4, got {ranks}"
        rank_ecpm = dict(zip(ranks, ecpms))
        assert rank_ecpm[1] >= rank_ecpm[2] >= rank_ecpm[3] >= rank_ecpm[4], (
            "Ranks not ordered by eCPM descending"
        )

    def test_yield_values(self):
        """eCPM in reasonable range, fill rate valid, and total revenue close to SSP total."""
        rows = self._rows()
        g = gt()
        for r in rows:
            ecpm = float(_g(r, "ecpm", 0) or 0)
            assert 1.0 <= ecpm <= 30.0, f"ecpm {ecpm} out of range for {_g(r, 'ssp')}"
            fr = float(_g(r, "fill_rate", 0) or 0)
            if fr > 2.0:
                fr = fr / 100.0
            assert 0 < fr <= 1.5, f"fill_rate {fr} out of range for {_g(r, 'ssp')}"
        total = sum(float(_g(r, "total_net_revenue_usd", 0) or 0) for r in rows)
        assert abs(total - g["total_ssp"]) / g["total_ssp"] < 0.05, (
            f"Yield total {total:.2f} vs SSP total {g['total_ssp']:.2f}"
        )


class TestVarianceAndSummary:

    def _variance_rows(self):
        wb = load_workbook(XLSX_OUT, data_only=True)
        _, rows = _sheet_rows(wb, "Variance_vs_Prior_Month")
        return rows

    def _exec_ws(self):
        wb = load_workbook(XLSX_OUT, data_only=True)
        ws = _find_sheet(wb, "Executive_Summary")
        assert ws is not None, "Executive_Summary sheet not found"
        return ws

    def test_variance_structure(self):
        """5 metric rows with correct names, valid statuses, and accurate prior-month values."""
        rows = self._variance_rows()
        g = gt()
        assert len(rows) == 5, f"Expected 5 metric rows, got {len(rows)}"
        expected_metrics = {
            "total_net_revenue_usd", "ivt_rate_pct", "discrepancy_rate_pct",
            "held_row_count", "avg_ecpm_usd",
        }
        actual_metrics = {str(_g(r, "metric", "")).lower().strip() for r in rows}
        assert expected_metrics.issubset(actual_metrics), (
            f"Missing metrics: {expected_metrics - actual_metrics}"
        )
        for r in rows:
            s = str(_g(r, "status", "")).lower()
            assert s in {"normal", "watch", "alert"}, f"Invalid variance status '{s}'"
            m = str(_g(r, "metric", "")).lower().strip()
            if m in g["variance_metrics"]:
                prior_exp = g["variance_metrics"][m]["prior"]
                prior_act = float(_g(r, "prior_month", 0) or 0)
                tol = max(0.1, abs(prior_exp) * 0.01)
                assert abs(prior_act - prior_exp) < tol, (
                    f"{m}: prior expected {prior_exp}, got {prior_act}"
                )

    def test_variance_change_pct_revenue(self):
        """change_pct is internally consistent for total_net_revenue (percentage, not fraction)."""
        for r in self._variance_rows():
            m = str(_g(r, "metric", "")).lower().strip()
            if m == "total_net_revenue_usd":
                cur = float(_g(r, "current_month", 0) or 0)
                pri = float(_g(r, "prior_month", 0) or 0)
                if pri > 0:
                    exp_chg = (cur - pri) / pri * 100
                    act_chg = float(_g(r, "change_pct", 0) or 0)
                    assert abs(act_chg - exp_chg) < 1.0, (
                        f"Revenue change_pct: expected {exp_chg:.1f}, got {act_chg:.1f}"
                    )

    def test_executive_summary(self):
        """Health indicator, makegood exposure, and top concerns are present and correct."""
        ws = self._exec_ws()
        g = gt()
        health = str(ws["B1"].value or "").strip()
        assert health.lower() in {"green", "yellow", "red"}, (
            f"overall_health should be Green/Yellow/Red, got '{health}'"
        )
        assert health.lower() == g["overall_health"].lower(), (
            f"overall_health expected '{g['overall_health']}', got '{health}'"
        )
        val = float(ws["B2"].value or 0)
        exp = g["total_makegood_dollars"]
        tol = max(1.0, exp * 0.05)
        assert abs(val - exp) < tol, (
            f"makegood_exposure {val:.2f} vs expected {exp:.2f}"
        )
        concerns = [
            row_idx for row_idx in range(6, 9)
            if ws.cell(row=row_idx, column=1).value is not None
        ]
        assert len(concerns) >= 1, "Expected at least 1 top concern in rows 6-8"
        for row_idx in concerns:
            src = str(ws.cell(row=row_idx, column=4).value or "")
            assert len(src) > 0, f"Row {row_idx} missing source_sheet reference"
