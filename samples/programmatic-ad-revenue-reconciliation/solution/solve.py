#!/usr/bin/env python3
"""Oracle solver: produces 5-sheet Excel audit_report.xlsx."""

import csv
import json
import os
import re
from collections import defaultdict
from datetime import date

import pdfplumber
from openpyxl import Workbook
from rapidfuzz import fuzz

DATA = "/root/data"
MONTH_START = date(2026, 4, 1)
REPORT_DATE = date(2026, 4, 30)
FUZZY_THRESH = 80
FLAG_USD = 5.0
FLAG_PCT = 3.0


def read_csv(path):
    with open(path) as f:
        return list(csv.DictReader(f))


def parse_date(s):
    s = s.strip()
    if re.match(r"\d{4}-\d{2}-\d{2}", s):
        return date.fromisoformat(s)
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
    return date.fromisoformat(s)


def read_pdf_billing(path):
    rows = []
    cols = ["ssp_line_id","month","deal_id","ssp_ad_unit_name","billed_impressions",
            "gross_revenue_usd","ssp_fee_pct","ssp_fee_usd","net_revenue_usd"]
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.split("\n"):
                if "|" not in line:
                    continue
                parts = [p.strip() for p in line.split("|")]
                if len(parts) < 5 or parts[0].startswith("-"):
                    continue
                if re.match(r"[A-Z]+-LINE-\d+", parts[0]):
                    row = {c: parts[i] if i < len(parts) else "" for i, c in enumerate(cols)}
                    rows.append(row)
    return rows


SSP_MAP = {"adflow_exchange": "adflow", "magnite": "magnite",
           "indexexchange": "indexexchange", "pubmatic": "pubmatic"}
SSP_LABEL = {"adflow": "AdFlow Exchange", "magnite": "Magnite",
             "indexexchange": "Index Exchange", "pubmatic": "Pubmatic"}


def load_ssp_billing():
    ssp_dir = os.path.join(DATA, "ssp_billing")
    rows = []
    for fname in sorted(os.listdir(ssp_dir)):
        ssp_id = None
        for prefix, sid in SSP_MAP.items():
            if fname.startswith(prefix):
                ssp_id = sid
                break
        if ssp_id is None:
            ssp_id = fname.split("_")[0].lower()
        fpath = os.path.join(ssp_dir, fname)
        if fname.endswith(".csv"):
            for r in read_csv(fpath):
                r["_ssp_id"] = ssp_id
                rows.append(r)
        elif fname.endswith(".pdf"):
            for r in read_pdf_billing(fpath):
                r["_ssp_id"] = ssp_id
                rows.append(r)
    return rows


def build_name_index(metadata):
    exact, names = {}, []
    for m in metadata:
        for sid in SSP_MAP.values():
            col = f"{sid}_name"
            if col in m and m[col]:
                exact[(sid, m[col])] = (m["ad_unit_id"], m["active"] == "true")
                names.append((sid, m[col], m["ad_unit_id"], m["active"] == "true"))
        exact[("_can", m["canonical_name"])] = (m["ad_unit_id"], m["active"] == "true")
        names.append(("_can", m["canonical_name"], m["ad_unit_id"], m["active"] == "true"))
    return exact, names


def resolve(ssp_id, name, exact, names):
    r = exact.get((ssp_id, name)) or exact.get(("_can", name))
    if r:
        return r
    best, bm = 0, None
    for sid, n, auid, act in names:
        if sid != ssp_id and sid != "_can":
            continue
        s = fuzz.ratio(name.lower(), n.lower())
        if s > best:
            best, bm = s, (auid, act)
    return bm if best >= FUZZY_THRESH else None


def main():
    ad_server = read_csv(os.path.join(DATA, "ad_server_delivery.csv"))
    ivt_data = read_csv(os.path.join(DATA, "ivt_report.csv"))
    rate_card = read_csv(os.path.join(DATA, "deal_rate_card.csv"))
    metadata = read_csv(os.path.join(DATA, "ad_unit_metadata.csv"))
    contract_goals = read_csv(os.path.join(DATA, "deal_contract_goals.csv"))
    request_log = read_csv(os.path.join(DATA, "ssp_request_log.csv"))
    with open(os.path.join(DATA, "historical_metrics.json")) as f:
        hist = json.load(f)
    ssp_rows = load_ssp_billing()

    au_lookup = {m["ad_unit_id"]: m for m in metadata}
    known_deals = set(rc["deal_id"] for rc in rate_card)
    exact, names = build_name_index(metadata)

    rc_lookup, rc_type = {}, {}
    for rc in rate_card:
        rc_lookup[(rc["deal_id"], rc["ad_unit_id"], rc["device"], rc["geo"])] = float(rc["cpm_usd"])
        rc_type[(rc["deal_id"], rc["ad_unit_id"])] = rc["rate_type"]

    # Delivery aggregation
    del_dpdg = defaultdict(int)
    for r in ad_server:
        del_dpdg[(r["deal_id"], r["ad_unit_id"], r["device"], r["geo"])] += int(r["delivered_impressions"])
    del_dp = defaultdict(int)
    for (d, au, dev, geo), v in del_dpdg.items():
        del_dp[(d, au)] += v

    # IVT (handles mixed date formats)
    ivt_dp = defaultdict(int)
    for r in ivt_data:
        ivt_dp[(r["deal_id"], r["ad_unit_id"])] += int(r["ivt_impressions"])

    # Ground truth per deal/AU
    dp_combos = sorted(del_dp.keys())
    gt = {}
    for dp in dp_combos:
        if dp not in rc_type:
            continue
        d, au = dp
        total_del = del_dp[dp]
        total_ivt = ivt_dp.get(dp, 0)
        billable = total_del - total_ivt
        wsum = sum(del_dpdg[(dd, aa, dv, ge)] * rc_lookup.get((dd, aa, dv, ge), 0)
                   for (dd, aa, dv, ge) in del_dpdg if dd == d and aa == au)
        wcpm = wsum / total_del if total_del > 0 else 0
        gt[dp] = {"del": total_del, "ivt": total_ivt, "bill": billable,
                  "cpm": wcpm, "rev": round(billable / 1000 * wcpm, 2),
                  "rt": rc_type[dp]}

    # ===== SHEET 1: Revenue Reconciliation =====
    total_exp = round(sum(g["rev"] for g in gt.values()), 2)
    total_ssp = round(sum(float(r["net_revenue_usd"]) for r in ssp_rows), 2)

    recon_rows, flagged, held, matched = [], [], [], set()
    for sr in ssp_rows:
        sid = sr["_ssp_id"]
        did = sr["deal_id"].strip()
        auname = sr["ssp_ad_unit_name"].strip()
        label = SSP_LABEL.get(sid, sid)

        def hold(reason):
            held.append({"ssp_line_id": sr["ssp_line_id"].strip(), "ssp": label, "reason": reason})
            recon_rows.append({"deal_id": did, "ad_unit_id": "", "canonical_name": "", "ssp": label,
                "delivered_impressions": 0, "ivt_impressions": 0, "billable_impressions": 0,
                "cpm_usd": 0, "expected_revenue_usd": 0,
                "ssp_billed_impressions": int(float(sr["billed_impressions"])),
                "ssp_net_revenue_usd": float(sr["net_revenue_usd"]),
                "discrepancy_usd": 0, "discrepancy_pct": 0, "status": "held_unmatched"})

        if did not in known_deals:
            hold("unknown_deal"); continue
        ai = resolve(sid, auname, exact, names)
        if ai is None:
            hold("unmatched_ad_unit"); continue
        auid, active = ai
        if not active:
            hold("inactive_ad_unit"); continue
        dp = (did, auid)
        if dp not in gt:
            hold("no_delivery_data"); continue

        matched.add(dp)
        g = gt[dp]
        snet = float(sr["net_revenue_usd"])
        disc = round(g["rev"] - snet, 2)
        pct = round(abs(disc) / g["rev"] * 100, 2) if g["rev"] > 0 else 0

        if g["rt"] == "floor" and disc < 0:
            status = "reconciled_above_floor"
        elif abs(disc) > FLAG_USD and pct > FLAG_PCT:
            status = "flagged_underbilled" if disc > 0 else "flagged_overbilled"
            flagged.append({"deal_id": did, "ad_unit_id": auid, "ssp": label,
                "reason": status, "expected_revenue_usd": g["rev"],
                "ssp_reported_revenue_usd": snet, "discrepancy_usd": disc, "discrepancy_pct": pct})
        else:
            status = "reconciled"

        recon_rows.append({"deal_id": did, "ad_unit_id": auid,
            "canonical_name": au_lookup[auid]["canonical_name"], "ssp": label,
            "delivered_impressions": g["del"], "ivt_impressions": g["ivt"],
            "billable_impressions": g["bill"], "cpm_usd": round(g["cpm"], 4),
            "expected_revenue_usd": g["rev"],
            "ssp_billed_impressions": int(float(sr["billed_impressions"])),
            "ssp_net_revenue_usd": snet, "discrepancy_usd": disc,
            "discrepancy_pct": pct, "status": status})

    total_adj = round(sum(gt[dp]["rev"] for dp in matched), 2)

    # ===== SHEET 2: Pacing & Make-Goods =====
    cg_lookup = {}
    for cg in contract_goals:
        cg_lookup[(cg["deal_id"], cg["ad_unit_id"])] = cg

    pacing_rows = []
    total_makegood = 0
    for cg in contract_goals:
        dp = (cg["deal_id"], cg["ad_unit_id"])
        if dp not in gt:
            continue
        g = gt[dp]
        contracted = int(cg["contracted_impressions"])
        cs = date.fromisoformat(cg["contract_start"])
        ce = date.fromisoformat(cg["contract_end"])
        days_total = (ce - cs).days + 1
        el_start = max(cs, MONTH_START)
        el_end = min(ce, REPORT_DATE)
        days_elapsed = max(1, (el_end - el_start).days + 1)
        delivered = g["del"]
        projected = int(delivered * days_total / days_elapsed)
        delivery_pct = round(delivered / contracted * 100, 2) if contracted > 0 else 0
        expected_pct = days_elapsed / days_total * 100
        mg_imps = max(0, contracted - projected)
        mg_dollars = round(mg_imps / 1000 * g["cpm"], 2)

        if delivery_pct >= 100:
            status = "complete"
        elif delivery_pct > expected_pct * 1.05:
            status = "ahead"
        elif delivery_pct >= expected_pct * 0.95:
            status = "on_track"
        elif delivery_pct >= expected_pct * 0.85:
            status = "at_risk"
        else:
            status = "behind"

        total_makegood += mg_dollars
        pacing_rows.append({
            "deal_id": cg["deal_id"], "ad_unit_id": cg["ad_unit_id"],
            "contracted_impressions": contracted,
            "contract_start": cg["contract_start"], "contract_end": cg["contract_end"],
            "days_elapsed": days_elapsed, "days_total": days_total,
            "delivered_impressions": delivered, "projected_eom_impressions": projected,
            "delivery_pct": delivery_pct, "makegood_impressions": mg_imps,
            "makegood_dollars": mg_dollars, "status": status})

    total_makegood = round(total_makegood, 2)

    # ===== SHEET 3: SSP Yield Analysis =====
    ssp_requests = defaultdict(int)
    for r in request_log:
        sid = r["ssp_id"]
        ssp_requests[sid] += int(r["ad_requests"])

    ssp_filled = defaultdict(int)
    ssp_revenue = defaultdict(float)
    ssp_flagged_ct = defaultdict(int)
    ssp_matched_ct = defaultdict(int)
    for sr in ssp_rows:
        sid = sr["_ssp_id"]
        ssp_filled[sid] += int(float(sr["billed_impressions"]))
        ssp_revenue[sid] += float(sr["net_revenue_usd"])
    for f in flagged:
        for sid, label in SSP_LABEL.items():
            if f["ssp"] == label:
                ssp_flagged_ct[sid] += 1
    for rr in recon_rows:
        if rr["status"] != "held_unmatched":
            for sid, label in SSP_LABEL.items():
                if rr["ssp"] == label:
                    ssp_matched_ct[sid] += 1

    yield_rows = []
    for sid in sorted(SSP_MAP.values()):
        filled = ssp_filled[sid]
        rev = round(ssp_revenue[sid], 2)
        reqs = ssp_requests[sid]
        ecpm = round(rev / filled * 1000, 2) if filled > 0 else 0
        fill = round(filled / reqs, 6) if reqs > 0 else 0
        mc = ssp_matched_ct[sid]
        dr = round(ssp_flagged_ct[sid] / mc, 4) if mc > 0 else 0
        yield_rows.append({"ssp": SSP_LABEL[sid], "ad_requests": reqs,
            "impressions_filled": filled, "fill_rate": fill,
            "total_net_revenue_usd": rev, "ecpm": ecpm,
            "discrepancy_rate": dr, "ranked_position": 0})

    yield_rows.sort(key=lambda r: -r["ecpm"])
    for i, yr in enumerate(yield_rows):
        yr["ranked_position"] = i + 1

    # ===== SHEET 4: Variance vs Prior Month =====
    total_delivered_all = sum(g["del"] for g in gt.values())
    total_ivt_all = sum(g["ivt"] for g in gt.values())
    cur_ivt_rate = round(total_ivt_all / total_delivered_all * 100, 2) if total_delivered_all > 0 else 0
    total_matched_rows = sum(1 for rr in recon_rows if rr["status"] != "held_unmatched")
    cur_disc_rate = round(len(flagged) / total_matched_rows * 100, 2) if total_matched_rows > 0 else 0
    all_filled = sum(ssp_filled.values())
    cur_ecpm = round(total_ssp / all_filled * 1000, 2) if all_filled > 0 else 0

    def var_status(pct):
        if abs(pct) <= 10:
            return "normal"
        if abs(pct) <= 25:
            return "watch"
        return "alert"

    def chg(cur, prior):
        return round((cur - prior) / prior * 100, 2) if prior != 0 else 0

    variance_rows = [
        {"metric": "total_net_revenue_usd", "current_month": total_ssp,
         "prior_month": hist["total_net_revenue_usd"],
         "change_pct": chg(total_ssp, hist["total_net_revenue_usd"]),
         "status": var_status(chg(total_ssp, hist["total_net_revenue_usd"]))},
        {"metric": "ivt_rate_pct", "current_month": cur_ivt_rate,
         "prior_month": hist["ivt_rate_pct"],
         "change_pct": chg(cur_ivt_rate, hist["ivt_rate_pct"]),
         "status": var_status(chg(cur_ivt_rate, hist["ivt_rate_pct"]))},
        {"metric": "discrepancy_rate_pct", "current_month": cur_disc_rate,
         "prior_month": hist["discrepancy_rate_pct"],
         "change_pct": chg(cur_disc_rate, hist["discrepancy_rate_pct"]),
         "status": var_status(chg(cur_disc_rate, hist["discrepancy_rate_pct"]))},
        {"metric": "held_row_count", "current_month": len(held),
         "prior_month": hist["held_row_count"],
         "change_pct": chg(len(held), hist["held_row_count"]),
         "status": var_status(chg(len(held), hist["held_row_count"]))},
        {"metric": "avg_ecpm_usd", "current_month": cur_ecpm,
         "prior_month": hist["avg_ecpm_usd"],
         "change_pct": chg(cur_ecpm, hist["avg_ecpm_usd"]),
         "status": var_status(chg(cur_ecpm, hist["avg_ecpm_usd"]))},
    ]

    # ===== SHEET 5: Executive Summary =====
    flagged_risk = round(sum(abs(f["discrepancy_usd"]) for f in flagged), 2)
    n_flagged = len(flagged)

    if n_flagged >= 3 or total_makegood > 2000:
        health = "Red"
    elif n_flagged >= 1 or total_makegood >= 500:
        health = "Yellow"
    else:
        health = "Green"

    concerns = []
    for f in flagged:
        concerns.append({
            "concern": f"Flagged {f['reason']}: {f['deal_id']}/{f['ad_unit_id']} via {f['ssp']}",
            "financial_impact_usd": round(abs(f["discrepancy_usd"]), 2),
            "source_sheet": "Revenue_Reconciliation"})
    for p in pacing_rows:
        if p["status"] in ("behind", "at_risk") and p["makegood_dollars"] > 0:
            concerns.append({
                "concern": f"Pacing {p['status']}: {p['deal_id']}/{p['ad_unit_id']} makegood needed",
                "financial_impact_usd": p["makegood_dollars"],
                "source_sheet": "Pacing_And_MakeGoods"})
    for vr in variance_rows:
        if vr["status"] == "alert":
            concerns.append({
                "concern": f"Variance alert: {vr['metric']} changed {vr['change_pct']:.1f}%",
                "financial_impact_usd": abs(vr["current_month"] - vr["prior_month"]) if isinstance(vr["current_month"], (int, float)) else 0,
                "source_sheet": "Variance_vs_Prior_Month"})
    concerns.sort(key=lambda c: -c["financial_impact_usd"])
    top3 = concerns[:3]

    # ===== Write Excel =====
    wb = Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Revenue_Reconciliation"
    h1 = ["deal_id","ad_unit_id","canonical_name","ssp","delivered_impressions",
          "ivt_impressions","billable_impressions","cpm_usd","expected_revenue_usd",
          "ssp_billed_impressions","ssp_net_revenue_usd","discrepancy_usd","discrepancy_pct","status"]
    ws1.append(h1)
    for rr in recon_rows:
        ws1.append([rr[k] for k in h1])

    # Sheet 2
    ws2 = wb.create_sheet("Pacing_And_MakeGoods")
    h2 = ["deal_id","ad_unit_id","contracted_impressions","contract_start","contract_end",
          "days_elapsed","days_total","delivered_impressions","projected_eom_impressions",
          "delivery_pct","makegood_impressions","makegood_dollars","status"]
    ws2.append(h2)
    for pr in pacing_rows:
        ws2.append([pr[k] for k in h2])

    # Sheet 3
    ws3 = wb.create_sheet("SSP_Yield_Analysis")
    h3 = ["ssp","ad_requests","impressions_filled","fill_rate","total_net_revenue_usd",
          "ecpm","discrepancy_rate","ranked_position"]
    ws3.append(h3)
    for yr in yield_rows:
        ws3.append([yr[k] for k in h3])

    # Sheet 4
    ws4 = wb.create_sheet("Variance_vs_Prior_Month")
    h4 = ["metric","current_month","prior_month","change_pct","status"]
    ws4.append(h4)
    for vr in variance_rows:
        ws4.append([vr[k] for k in h4])

    # Sheet 5
    ws5 = wb.create_sheet("Executive_Summary")
    ws5["A1"] = "overall_health"
    ws5["B1"] = health
    ws5["A2"] = "total_makegood_exposure_usd"
    ws5["B2"] = total_makegood
    ws5["A3"] = "flagged_revenue_at_risk_usd"
    ws5["B3"] = flagged_risk
    ws5.append([])  # row 4 blank
    ws5.append(["rank", "concern", "financial_impact_usd", "source_sheet"])
    for i, c in enumerate(top3):
        ws5.append([i + 1, c["concern"], c["financial_impact_usd"], c["source_sheet"]])

    wb.save("/root/audit_report.xlsx")
    print(f"Done. Wrote /root/audit_report.xlsx with {len(recon_rows)} recon rows, "
          f"{len(pacing_rows)} pacing rows, {len(yield_rows)} yield rows, "
          f"{len(flagged)} flagged, {len(held)} held, health={health}")


if __name__ == "__main__":
    main()
